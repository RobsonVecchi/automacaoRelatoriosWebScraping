from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from time import sleep
import datetime
from openpyxl import Workbook, load_workbook

def start_driver():
    chrome_options = Options()

    arguments = ['--lang=pt-BR', 'window-size=800,600', '--incognito', '--headless']
    for argument in arguments:
        chrome_options.add_argument(argument)
        
    driver = webdriver.Chrome(options=chrome_options)   
    return driver 

def format_price(price):
    price = price.strip()
    price = price.replace(".", "")
    price = price.replace("R$", "").replace(",", ".")
    price = float(price)
    final_price = f"{price:.2f}"
    return final_price

def collect_data():
    driver.get('https://www.norrisimports.com.br/produto/guitarra-6-cordas-s-by-solar-eb46w-branca-fosca-explorer.html')
    driver.implicitly_wait(10)
    price_element = driver.find_element(By.XPATH, '//strong[@class="cor-principal titulo"]')
    pure_price = price_element.text
    final_price = format_price(pure_price)
    driver.quit()
    return final_price

def save_to_excel():
    product_name = 'Guitarra 6 Cordas S by Solar EB4.6W branca fosca Explorer'
    formatted_price = collect_data()
    product_link = 'https://www.norrisimports.com.br/produto/guitarra-6-cordas-s-by-solar-eb46w-branca-fosca-explorer.html'
    current_date = datetime.date.today()
    excel_filename = 'product_data.xlsx'
    
    try:
        # Try to load the existing workbook
        wb = load_workbook(excel_filename)
        ws = wb.active
        
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        wb = Workbook()
        ws = wb.active
        # Set column headers for a new file
        ws['A1'] = 'Product'
        ws['B1'] = 'Date'
        ws['C1'] = 'Price'
        ws['D1'] = 'Link'

    #Update and save excel file    
    ws.append([product_name, current_date, formatted_price, product_link])
    wb.save(excel_filename)

driver = start_driver()
    
while True:
    save_to_excel()
    sleep(1800)
